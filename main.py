import requests
import json
import datetime
import openpyxl
import pandas
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


################################################################################################################

today = datetime.date.today()       # Get today's date in YYYY-MM-DD format

keys = []
keys_w_add_drop = []
current_row = 2

cdns = {
'154' : 'Xinnet',
'157' : 'Cloudflare CDN',
'204' : 'EdgeCast',
'340' : 'Imperva Incapsula',
'378' : 'CDN77',
'400' : 'Akamai',
'459' : 'Cloudflare CDN',
'494' : 'CDNetworks',
'496' : 'Fastly',
'497' : 'Azure',
'500' : 'MaxCDN',
'501' : 'Amazon CloudFront',
'513' : 'Akamai',
'515' : 'BitGravity',
'6576' : 'Cloudinary',
'6828' : 'Limelight',
'21136' : 'KeyCDN',
'35468' : 'Stackpath',
'38622' : 'RawGit',
'44802' : 'JsDelivr',
'6547' : 'OnApp CDN'
}



datanyze_url = 'https://api.datanyze.com/alerts/?email=florian.parzhuber@cdnetworks.com&token=25c368f5afa267a1b31870fe8cda3c09&date=' + str(today)

url = requests.get(datanyze_url)
jsons = json.loads(url.text)


### FIND NUMBER OF ADDS AND DROP FOR A PARTICULAR CDN ###
for key in jsons:
    keys.append(key)

for i in range(len(keys)):
    keys_w_add_drop.append(keys[i])
    try:
        number_adds = len(jsons[keys[i]]["1"])
        keys_w_add_drop.append(number_adds)
    except KeyError:
        number_adds = 0
        keys_w_add_drop.append(number_adds)

    try:
        number_drops = len(jsons[keys[i]]["2"])
        keys_w_add_drop.append(number_drops)
    except KeyError:
        number_drops = 0
        keys_w_add_drop.append(number_drops)


#print (keys_w_add_drop)

### SETTING UP THE EXCEL SPREADSHEET ###

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row = 1, column = 2).value ="Domain"
ws.cell(row = 1, column = 2).font= Font(bold=True)
ws.cell(row = 1, column = 3).value ="Alexa Rank"
ws.cell(row = 1, column = 3).font= Font(bold=True)
ws.cell(row = 1, column = 3).alignment = Alignment(horizontal='center')
ws.cell(row = 1, column = 4).value ="CDN Technology"
ws.cell(row = 1, column = 4).font= Font(bold=True)
ws.cell(row = 1, column = 4).alignment = Alignment(horizontal='center')
ws.cell(row = 1, column = 5).value ="Add / Drop"
ws.cell(row = 1, column = 5).font= Font(bold=True)
ws.cell(row = 1, column = 5).alignment = Alignment(horizontal='center')

### NOW WE WILL DO DE ACTUAL WRITING TO AN EXCEL SPREADSHEET ###

key_number = int(len(keys_w_add_drop))

for i in range(0,key_number,3):

    for x in range(keys_w_add_drop[i+1]):
        ws.cell(row = len(ws['B'])+1, column = 2).value = jsons[keys_w_add_drop[i]]["1"][x]["domain"]
        
        if int(jsons[keys_w_add_drop[i]]["1"][x]["alexa_rank"]) == 1000001:
            ws.cell(row = len(ws['C']), column = 3).value = 16776089
        else:
            ws.cell(row = len(ws['C']), column = 3).value = jsons[keys_w_add_drop[i]]["1"][x]["alexa_rank"]

        ws.cell(row = len(ws['C']), column = 3).alignment = Alignment(horizontal='center')
        ws.cell(row = len(ws['D']), column = 4).value = cdns[keys_w_add_drop[i]]
        ws.cell(row = len(ws['D']), column = 4).alignment = Alignment(horizontal='center')
        ws.cell(row = len(ws['E']), column = 5).value = "Add"
        ws.cell(row = len(ws['E']), column = 5).alignment = Alignment(horizontal='center')

for i in range(0,key_number,3):
    for z in range(keys_w_add_drop[i+2]):
        ws.cell(row = len(ws['B'])+1, column = 2).value = jsons[keys_w_add_drop[i]]["2"][z]["domain"]

        if int(jsons[keys_w_add_drop[i]]["2"][z]["alexa_rank"]) == 1000001:
            ws.cell(row = len(ws['C']), column = 3).value = 16776089
        else:
            ws.cell(row = len(ws['C']), column = 3).value = jsons[keys_w_add_drop[i]]["2"][z]["alexa_rank"]

        ws.cell(row = len(ws['C']), column = 3).alignment = Alignment(horizontal='center')
        ws.cell(row = len(ws['D']), column = 4).value = cdns[keys_w_add_drop[i]]
        ws.cell(row = len(ws['D']), column = 4).alignment = Alignment(horizontal='center')
        ws.cell(row = len(ws['E']), column = 5).value = "Drop"
        ws.cell(row = len(ws['E']), column = 5).alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 10

filter_range = 'B1:E' + str(len(ws['B']))
ws.auto_filter.ref = filter_range

wb.save("C:/Users/User/Desktop/Daily Datanyze Export/" + str(today) + '.xlsx') # Save data to worksheet with today's date as file name











